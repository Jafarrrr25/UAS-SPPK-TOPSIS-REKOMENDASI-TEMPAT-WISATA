from flask import Flask, request, render_template
import pandas as pd
import numpy as np
import openpyxl

app = Flask(__name__)

data = pd.read_excel('Dataset Wisata.xlsx')
alternatif = data['Nama Tempat']
kriteria = ['Rating Maps', 'Harga Tiket', 'Komentar', 'Respon Pemilik', 'Foto']
costbenefit = ["benefit", "benefit", "benefit", "benefit", "benefit"]
kepentingan = [5,5,4,3,5]
alternatifKriteria = data.iloc[:,[2,3,4,5,6]]

@app.route('/')
def home():
    return render_template('index.html')

@app.route('/lihat_data')
def lihat_data():
    # Jalankan kode program Python Anda di sini
    data = pd.read_excel('Dataset Wisata.xlsx')  # misalnya Anda membaca file CSV
    data = data.drop('No', axis=1)

    return render_template('lihat_data.html', tables=[data.to_html(classes='data')], titles=data.columns.values)

@app.route('/lihat_pembobotan')
def lihat_pembobotan():
    # Jalankan kode program Python Anda di sini
    kriteria = pd.read_excel('Kriteria.xlsx')
    
    return render_template('lihat_pembobotan.html', tables=[kriteria.to_html(classes='data')], titles=kriteria.columns.values)

@app.route('/tambah_data', methods=['GET', 'POST'])
def tambah_data():
    if request.method == 'POST':
        # Buka file excel
        wb = openpyxl.load_workbook('Dataset Wisata.xlsx')
        sheet = wb.active

        # cari baris terakhir untuk menentukan nilai "No" yang baru
        last_row_no = sheet.max_row

        # Hitung nilai No yang baru
        new_no = last_row_no + 1

        nama_tempat = request.form['nama_tempat']
        rating_maps = request.form['rating_maps']
        harga_tiket = request.form['harga_tiket']
        komentar = request.form['komentar']
        respon_pemilik = request.form['respon_pemilik']
        foto = request.form['foto']

        # Tambahkan data ke excelnya
        new_data = [new_no, nama_tempat, rating_maps, harga_tiket, komentar, respon_pemilik, foto]
        sheet.append(new_data)

        # Simpan dan tutup 
        wb.save('Dataset Wisata.xlsx')
        wb.close()

    # Menampilkan excel 
    wb = openpyxl.load_workbook('Dataset Wisata.xlsx')
    sheet = wb.active
    data = list(sheet.iter_rows(min_row=2, values_only=True))

    # Tidak memasukkan kolom "No"
    data_without_no = [row[1:] for row in data]

    wb.close()

    return render_template('tambah_data.html', data=data_without_no)

@app.route('/topsis')
def topsis():
    # Jalankan kode program Python Anda di sini
    data = pd.read_excel('Dataset Wisata.xlsx')
    
    return render_template('topsis.html', tables=[data.to_html(classes='data')], titles=data.columns.values)

@app.route('/topsis/pembagi', methods=['POST'])
def pembagi():
    global matrik_pembagi
    #Mencari Pembaginya
    pembagi = []
    for i in range (len(kriteria)):
        pembagi.append(0)
        for j in range (len(alternatif)):
            pembagi [i] = pembagi[i] + (alternatifKriteria.iloc[j,i] * alternatifKriteria.iloc[j,i])
        pembagi[i] = np.sqrt(pembagi[i])
    matrik_pembagi = pd.DataFrame([pembagi], columns=['Rating Maps', 'Harga Tiket', 'Komentar', 'Respon Pemilik', 'Foto'])

    # Ubah DataFrame menjadi HTML dan kembalikan sebagai respons
    return matrik_pembagi.to_html()
    
@app.route('/topsis/normalisasi', methods=['POST'])
def normalisasi():
    global matrik_pembagi
    global normalisasi1
    #Normalisasi
    normalisasi = []
    for i in range (len(alternatif)):
        normalisasi.append([])
        for j in range (len(kriteria)):
            normalisasi[i].append(0)
            normalisasi[i][j] = alternatifKriteria.iloc[i,j]/matrik_pembagi.iloc[0,j]
    data.iloc[:, 2:7] = pd.DataFrame(normalisasi, columns=['Rating Maps', 'Harga Tiket', 'Komentar', 'Respon Pemilik', 'Foto'])
    normalisasi = data
    normalisasi1 = data.iloc[:, 2:7]

    # Ubah DataFrame menjadi HTML dan kembalikan sebagai respons
    return normalisasi.to_html()

@app.route('/topsis/ternormalisasi', methods=['POST'])
def ternormalisasi():
    global matrik_pembagi
    global normalisasi1
    global terbobot1
    #Matriks Ternormalisasi
    terbobot = []
    for i in range(len(alternatif)):
        terbobot.append([])
        for j in range (len(kriteria)):
            terbobot[i].append(0)
            terbobot[i][j] = normalisasi1.iloc[i,j] * kepentingan[j]
    data.iloc[:, 2:7] = pd.DataFrame(terbobot, columns=['Rating Maps', 'Harga Tiket', 'Komentar', 'Respon Pemilik', 'Foto'])
    terbobot = data
    terbobot1 = data.iloc[:, 2:7]

    # Ubah DataFrame menjadi HTML dan kembalikan sebagai respons
    return terbobot.to_html()

@app.route('/topsis/solusi/minus', methods=['POST'])
def minSolusi():
    global terbobot1
    global minusSolution

    #mencari Solusi A-
    minSolution = []
    for i in range (len(kriteria)):
        minSolution.append(0)
        if costbenefit[i] == 'cost' : 
            for j in range(len(alternatif)):
                if j == 0:
                    minSolution[i] = terbobot1.iloc[j,i]
                else: 
                    if minSolution[i] < terbobot1.iloc[j,i]:
                        minSolution[i] = terbobot1.iloc[j,i]
        else: #costbenefit[i] == 'benefit':
            for j in range(len(alternatif)):
                if j == 0:
                    minSolution[i] = terbobot1.iloc[j,i]
                else:
                    if minSolution[i] > terbobot1.iloc[j,i]:
                        minSolution[i] = terbobot1.iloc[j,i]
    minusSolution = pd.DataFrame(minSolution, columns=['Min_Solutions'])

    return minusSolution.to_html()

@app.route('/topsis/solusi/plus', methods=['POST'])
def plusSolusi():
    global terbobot1
    global plusSolution
    
    #mencari Solusi A+
    plusSolution = []
    for i in range (len(kriteria)):
        plusSolution.append(0)
        if costbenefit[i] == 'cost' : 
            for j in range(len(alternatif)):
                if j == 0:
                    plusSolution[i] = terbobot1.iloc[j,i]
                else: 
                    if plusSolution [i] > terbobot1.iloc[j,i]:
                        plusSolution[i] = terbobot1.iloc[j,i]
        else: #costbenefit[i] == 'benefit':
            for j in range(len(alternatif)):
                if j == 0:
                    plusSolution[i] = terbobot1.iloc[j,i]
                else:
                    if plusSolution[i] < terbobot1.iloc[j,i]:
                        plusSolution[i] = terbobot1.iloc[j,i]
    plusSolution = pd.DataFrame(plusSolution, columns=['Plus_Solutions'])

    return plusSolution.to_html()

@app.route('/topsis/jarak/plus', methods=['POST'])
def plusAlternatif():
    global terbobot1
    global plusSolution
    global plusAlternatif
    
   # menghitung Jarak Alternatif A+
    plusAlternate = []
    for i in range(len(alternatif)):
        plusAlternate.append(0)
        for j in range(len(kriteria)):
            plusAlternate[i] = plusAlternate[i] + ((plusSolution.iloc[j,0] - terbobot1.iloc[i,j]) * (plusSolution.iloc[j,0] - terbobot1.iloc[i,j]))
        plusAlternate[i] = plusAlternate[i]**(1/2.0)

    plusAlternatif = pd.DataFrame(plusAlternate, columns=['Plus_Alternate'])
    return plusAlternatif.to_html()

@app.route('/topsis/jarak/minus', methods=['POST'])
def minAlternatif():
    global terbobot1
    global minusSolution
    global minAlternatif
    
   # menghitung Jarak Alternatif A-
    minAlternate = []
    for i in range(len(alternatif)):
        minAlternate.append(0)
        for j in range(len(kriteria)):
            minAlternate[i] = minAlternate[i] + ((terbobot1.iloc[i,j] - minusSolution.iloc[j,0]) * (terbobot1.iloc[i,j] - minusSolution.iloc[j,0]))
        minAlternate[i] = minAlternate[i]**(1/2.0)
    minAlternatif = pd.DataFrame(minAlternate, columns=['Min_Alternate'])
    return minAlternatif.to_html()

@app.route('/topsis/preferensi', methods=['POST'])
def preferensi():
    global minAlternatif
    global plusAlternatif
    global preferensi
    #Hasil preferensi
    hasil = []
    for i in range(len(alternatif)):
        hasil.append(0)
        for j in range(len(kriteria)):
            hasil[i] = plusAlternatif.iloc[i,0] / (minAlternatif.iloc[i,0]+plusAlternatif.iloc[i,0])
    
    preferensi = pd.DataFrame(hasil, columns=['Nilai Preferensi'])

    return preferensi.to_html()

@app.route('/topsis/rank', methods=['POST'])
def rank():
    global preferensi
    #Ranking berdasarkan preferensi
    alterRank = []
    resultRank = []
    for i in range(len(alternatif)):
        resultRank.append(preferensi.iloc[i,0])
        alterRank.append(alternatif[i])

    for i in range(len(alternatif)):
        for j in range(len(alternatif)):
            if j > i:
                if resultRank[j] > resultRank[i]:
                    tmphasil = resultRank[i]
                    tmpalternatif = alterRank[i]
                    resultRank[i] = resultRank[j]
                    alterRank[i] = alterRank[j]
                    resultRank[j] = tmphasil
                    alterRank[j] = tmpalternatif
    rank = pd.DataFrame({'Nama Tempat': alterRank, 'Ranking Preferensi': resultRank})

    return rank.to_html()

if __name__ == '__main__':
    app.run(debug=True)